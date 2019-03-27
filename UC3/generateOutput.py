from bs4 import BeautifulSoup
import json
from openpyxl import load_workbook
from urllib.parse import unquote
import pandas as pd
import os
import config
import numpy as np
import re
import argparse
import logging
import requests
from collections import OrderedDict
import time


def getJsonFiles(path):
    files = []
    try:
        for jFile in os.listdir(path):
            if jFile.endswith('.json'):
                jFile = os.path.join(path, jFile)
                files.append(jFile)
        files = sorted(files, key=os.path.getmtime)
        logger.info("Processing {} Json's".format(len(files)))
    except Exception as e:
        logger.error("Error while getting Json files \n Detail:{}".format(e))
    return files


def writeToExcel(outFile, sheets, idx=True, hdr=True):
    try:
        book = None
        writer = pd.ExcelWriter(outFile, engine='openpyxl')
        # if output file exists get other sheets data from there, create new file otherwise
        if os.path.exists(outFile):
            book = load_workbook(outFile)
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        for sheet in sheets.keys():
            df = sheets[sheet]
            df = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
            df.to_excel(writer, sheet_name=sheet, index=idx, header=hdr)

        writer.save()
        writer.close()
    except Exception as e:
        logger.error("Error while saving report\nDetail: {}".format(e))
    else:
        logger.info("Report saved successfully to file: {}".format(outFile))
    finally:
        pass


def getData(jsonFile, params):
    data = getFileData(jsonFile)
    filterList = config.filterList
    udata = {}
    entryUrl = ""
    url = ""
    try:
        url = data["log"]["pages"][0]["title"]
        for entry in data["log"]["entries"]:
            for filter in filterList:
                if filter in entry['request']['url']:
                    entryUrl = unquote(entry['request']['url'])
                    query_data = entry['request'].get('queryString', None)
                    query_data.extend(entry['request'].get('headers', None))
                    if query_data:
                        udata[filter + '|:path'] = entryUrl
                        for e in query_data:
                            if e['name'].strip() in params:
                                udata[e['name']] = unquote(e['value'])
                        logger.info("Found QueryData in {} for {} by applying this filter {}".format(jsonFile, url, filter))
                    else:
                        logger.warning("{} with {} has No required parameters".format(jsonFile, url))
                    break
        if not udata:
            logger.warn("{} with {} has No required parameters".format(jsonFile, url))
    except Exception as e:
        logger.error("Error occurred while retrieving query data from " + jsonFile + ".\ndetail: {}".format(e))
    return udata, entryUrl, url


def getFileData(file, raw=False):
    data = ''
    with open(file, encoding="utf8") as f:
        if raw:
            return f.read()
        try:
            data = json.load(f)
        except Exception as e:
            data = ""
            logger.error("Error while converting file data to json")
    return data


def getMasterData(xl):
    try:
        # df_Master = xl.parse(config.sheetname)
        df_Master = xl.parse(0)
        df_obj = df_Master.select_dtypes(['object'])
        df_Master[df_obj.columns] = df_obj.apply(lambda x: x.str.strip())
        df_Master.set_index('parameters', inplace=True)
        df_Master = df_Master.replace(np.nan, 'None', regex=True)
        return df_Master
    except Exception as e:
        logger.error("Error in getting SheetName of Master File " + ".\ndetail: {}".format(e))


def getReport_Suite(jsonFile, params):
    data = getFileData(jsonFile)
    url = ""
    prodUrl1 = None
    try:
        url = data["log"]["pages"][0]["title"]
        for entry in data["log"]["entries"]:
            if "/b/ss/" in entry['request']['url']:
                prodUrl = entry['request']['url']
                prodUrl1 = prodUrl.replace('/b/ss/', '|').split('|')[1].split('/')[0]
                logger.info("Found Report-Suite {}".format(prodUrl1))
                break
    except Exception as e:
        logger.warning("Error occurred while retrieving Report-Suite from " + jsonFile + ".\ndetail: {}".format(e))

    return prodUrl1, url


def getDTM_Header_Script(url):
    res = requests.get(url)
    soup = BeautifulSoup(res.text, "html.parser")
    items = soup.find_all("script", src=lambda value: value and value.startswith("//assets.adobedtm.com"))
    if items:
        assetdata = items[0]['src']
        logger.info("{} has DTM Header Script".format(url))
        return assetdata
    else:
        logger.warning("{} has no DTM Header Script".format(url))


def deleteOutputExcel(filename):
    if os.path.exists(filename):
        os.remove(filename)


def hideExcelSheet(filename):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    ws.sheet_state = 'hidden'
    wb.save(filename)


def color_negative_red(val):
    color = 'red' if val == "Fail" else 'black'
    return 'color: %s' % color


def highlight_max(s):
    '''
    '''
    is_max = s == "Fail"
    return ["background-color: yellow" if v else '' for v in is_max]


parser = argparse.ArgumentParser()
parser.add_argument('--inputExcelFilePath', help="Error in Input Excel file path", type=str, required=True)
parser.add_argument('--outputFilePath', help="Error Output Excel file path", type=str, required=True)
parser.add_argument('--exceptionFilePath', help="Error in Exception file path", type=str, required=True)
parser.add_argument('--jsonFilePath', help="Error in Json file path", type=str, required=True)
parser.add_argument('--logFilePath', help="Error in log file path", type=str, required=True)

parser.add_argument('--radioButton', help="Error in selecting Radio Button", type=str, required=False)


args = parser.parse_args()

inputExcelFilePath = args.inputExcelFilePath
outputFilePath = args.outputFilePath
exceptionFilePath = os.path.join(args.exceptionFilePath, "Exceptions.txt")
jsonFilePath = args.jsonFilePath
logFilePath = os.path.join(args.logFilePath, "Logs.txt")
radioButton = args.radioButton


logging.basicConfig(filename=logFilePath, level=logging.DEBUG, format="%(asctime)s : %(levelname)s : %(message)s")
logger = logging.getLogger()

logger.info("Python Process started")

json_files = getJsonFiles(jsonFilePath)

try:
    xl = pd.ExcelFile(inputExcelFilePath)
except Exception as e:
    logger.warning("Error in Reading Master Excel File \n Detail: {}".format(e))


if outputFilePath:
    logger.info("Deleting the existing output file")
    deleteOutputExcel(outputFilePath)
else:
    pass

try:
    df_Master = getMasterData(xl)
    params = list(df_Master.index)
    masterUrls = df_Master.columns.tolist()
    masterResult = OrderedDict()
    masterResult['parameters'] = params
    inputDf = pd.DataFrame(index=params)
except Exception as e:
    logger.error("Error Occurred while getting Master File , n Detail : {}".format(e))


urlProcessed = []
failResult = [['URL', 'Parameter', 'ActualValue', 'ExpectedValue']]
for inputjson in json_files:
    query_data = {}
    query_data, entryUrl, url = getData(inputjson, params)
    logger.info("Processing " + url + " with QueryData")
    logger.info("Processing " + url + " with Report_Suite ,DTM_Header_Script ")
    Report_Suite, url = getReport_Suite(inputjson, params)
    # url = "https://www.google.com/"
    url1 = url.split('|')[0]
    DTM_Header_Script = getDTM_Header_Script(url1)
    query_data['DTM_Header_Script'] = DTM_Header_Script
    query_data['Report_Suite'] = Report_Suite
    query_data[':path'] = entryUrl
    if url in urlProcessed:
        logger.warning("This {} is already processed".format(url))
        continue
    urlProcessed.append(url)
    logger.info(query_data.keys())
    if not query_data:
        continue
    df_Input = pd.DataFrame.from_dict(query_data, orient='index')
    df_Input.columns = [url]
    inputDf = inputDf.join(df_Input, how="left")
    inputDf.replace(np.nan, 'None', regex=True, inplace=True)
    logger.info("Processing " + inputjson + " with URL: " + url)
    urlResult = []
    if url in masterUrls:
        try:
            for p in params:
                r = " "
                masterV = str(df_Master.loc[p][url]).strip().lower()
                inputV = str(inputDf.loc[p][url]).strip().lower()
                if "none" not in masterV:
                    if inputV == masterV:
                        r = 'Pass'
                        urlResult.append(r)
                    else:
                        try:
                            # result = re.match(masterV, inputV, re.IGNORECASE)
                            result = re.search(masterV, inputV, re.IGNORECASE)
                            if result:
                                r = 'Pass'
                            else:
                                r = "Fail"
                                element = [url, p, inputV, masterV]
                                failResult.append(element)

                        except Exception as e:
                            r = "Fail"
                            element = [url, p, inputV, masterV]
                            failResult.append(element)
                            pass
                        urlResult.append(r)
                else:
                    r = "NA"
                    urlResult.append(r)
            masterResult[url] = urlResult
        except Exception as e:
            logger.warning("Detail: {}".format(e))

    else:
        logger.warning(url + " Not found in master sheet")
    logger.info("Processed " + inputjson + " with URL: " + url)
try:
    masterReport = pd.DataFrame(masterResult)
    masterReport.set_index('parameters', inplace=True)
    masterReport_T = masterReport.T
    masterReport_T.index.name = 'parameters'
    inputDf.index.name = 'parameters'
    masterReport_T = masterReport_T.style.applymap(color_negative_red).apply(highlight_max)
    inputDf = inputDf.reindex(params)
    writeToExcel(outputFilePath, {"Input": inputDf, "Output": masterReport_T})
    fHeaders = failResult.pop(0)
    failResultDf = pd.DataFrame(failResult, columns=fHeaders)
    writeToExcel(outputFilePath, {"FailedParameters": failResultDf}, idx=False)

    if os.path.exists(exceptionFilePath):
        exceptionData = getFileData(exceptionFilePath, raw=True).split('\n')
        exceptionDf = pd.DataFrame(exceptionData)
        writeToExcel(outputFilePath, {"Exceptions": exceptionDf}, idx=False, hdr=False)

    hideExcelSheet(outputFilePath)
    logger.info("Python Process is Completed")


except Exception as e:
    logger.error("Error \nDetail: {}".format(e))
